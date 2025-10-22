import streamlit as st  
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
st.set_page_config(layout="wide")
GRlogo = str(Path(__file__).parent/ "RGS_logo.png")
st.logo(GRlogo,size="large")
st.subheader("ICMM and INAP ARD/ML Decision Support System",divider = 'grey')
#st.write("--------------------------------")

link = '[INAP/ICMM](https://www.icmm.com/en-gb/guidance/environmental-stewardship/2025/ardml)'
st.markdown("Notes on the ARD/ML prevention and management tool: "+link, unsafe_allow_html=True)
# Load the core database of DSS records from the Excel workbook ("Database" sheet)
# Note: the optional display below is useful for debugging the raw data shape
#df=pd.read_excel('ICMM-INAP-ARD_ML-ToolUNLOCKED_Automatic Password setting removed.xlsm',sheet_name='Database')
#st.dataframe(df.drop(columns='DSS No.'),width=1000,hide_index=True)

df=pd.read_csv('database.csv')
#st.dataframe(df.head())
# Build the set of available Asset development stages from the database
Asset_dev_stages=df['Asset development stages'].unique()


# Build the set of available Leading practice activities from the database
Leading_prac_activity=df['Leading practice activity'].unique()




with st.container(width=900,horizontal_alignment="center"):
    col1, col2, col3 = st.columns([20,60,20])
    with col2:
        st.image("MatrixPic.png",width=800)
    col1, col2, col3 = st.columns([20,40,40])
    with col2:
    # UI control: select one Asset development stage
        selected_Asset_dev_stages=st.radio(
            options=Asset_dev_stages,
            key='Asset_dev_stages',
            label='Asset development stages'
        )
    available_options=df[(df['Asset development stages']==selected_Asset_dev_stages) & (df['Content']!= "N")]
    available_options=available_options['Leading practice activity'].unique()
    with col3:
        # UI control: select one Leading practice activity from the filtered list
        selected_Leading_prac_activity=st.radio(
            label='**Leading practice activity**',
            options=available_options,
            key='Leading_prac_activity'

            )



#st.divider()  # Visual separator for readability

    

#Todo make a dictionary of lists of image locations so that the images are displayed below the table.
# Narrow the database to the single record for the chosen stage and activity
df_filtered=df[df['Asset development stages']==selected_Asset_dev_stages]
df_filtered=df_filtered[df_filtered['Leading practice activity']==selected_Leading_prac_activity]

# Optional: preview the filtered row(s)
#st.table(df_filtered)

# The Leading practice area is shown as part of the generated support text
leading_practice_area=df_filtered['Leading practice area'].unique()


#Load the "Simple setup" sheet, which maps (activity, stage) to a DSS record number
# setup_number = pd.read_excel(
#     'ICMM-INAP-ARD_ML-ToolUNLOCKED_Automatic Password setting removed.xlsm',
#     sheet_name='Simple setup'
# )
setup_number = pd.read_csv('Simple_setup.csv')
setup_number.columns = setup_number.columns.str.strip()  # Guard against stray header whitespace


# Retrieve the DSS record number by selecting the row that matches the chosen
# activity, then taking the value under the chosen stage column
return_number = setup_number.loc[setup_number['Leading practice activity'] == selected_Leading_prac_activity,selected_Asset_dev_stages].iloc[0]


# Optional: inspect the selected DSS number
#st.write(return_number)

# Present a simple header section describing the generated support and selections
col1, col2, col3 = st.columns([15,80,10])
with col2:
    st.subheader("**Leading Practice Area:** "+ leading_practice_area[0])
    st.markdown("**Leading practice activity:** "+selected_Asset_dev_stages)
    st.markdown("**Asset development stage:** "+selected_Leading_prac_activity)
    

# Build the record to display: locate the row with the selected DSS number,
# drop metadata columns not intended for end-user display, and transpose so
# that aspects appear as rows with a single value column
display_df=df[df["DSS No."]==return_number].drop(columns=['Content','Asset development stages','Leading practice activity','Leading practice area','DSS No.']).T
display_df.index.name = 'Aspect Considered'
display_df.columns = ['Description']  # Blank column header for a cleaner table
path = 'Output data.xlsx'
display_df.to_excel(path, index=True, sheet_name="Sheet1")



wb = load_workbook(path)
ws = wb.active  # first sheet

# Set column widths
ws.column_dimensions['A'].width = 37
ws.column_dimensions['B'].width = 85

for row_num in range(1, ws.max_row + 1): # Iterate through all rows
    cell = ws.cell(row=row_num, column=2) # Access cell in column B
    cell.alignment = Alignment(wrapText=True)
wb.save(path)
wb.close()
        



col1, col2, col3 = st.columns([15,80,10])
with col2:
    with open('Output data.xlsx', "rb") as file:
        btn = st.download_button(
                label="Download as excel file",
                data=file.read(),
                file_name="Output data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
with col2:
    # CSS: wrap long text inside data editor cells
    st.markdown("""
    <style>
    /* Make the table use fixed layout so column widths apply */
    table {
        table-layout: fixed;
        width: 100%;
    }
    /* First column wider */
    thead th:first-child, tbody td:first-child {
        width: 10%;  /* adjust as needed */
    }
    /* Second column takes the rest */
    thead th:nth-child(2), tbody td:nth-child(2) {
        width: 65%;
    }
    /* Wrap long words and keep line breaks normal */
    th, td {
        white-space: normal !important;
        word-break: normal;
        overflow-wrap: anywhere;
        vertical-align: top;
    }
    </style>
    """, unsafe_allow_html=True)

    st.table(display_df) 





